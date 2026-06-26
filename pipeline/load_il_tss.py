#!/usr/bin/env python3
"""
Load Illinois Teacher Salary Study (TSS) data into tss_annual.
Upserts a districts row (state='IL') for each distinct district found.

Usage:
    python3 pipeline/load_il_tss.py
"""
import json
import logging
import re
import sys
from collections import namedtuple
from pathlib import Path
from typing import Any, Optional

sys.path.insert(0, str(Path(__file__).parent))
import common

common.setup_logging()
log = logging.getLogger(__name__)

IL_TSS_DIR = common.DATA_DIR / "il_tss"

# Ordered so chronological sort also works by school_year string.
FILE_SCHOOL_YEARS: dict[str, str] = {
    "teacher_salary_data_15-16.xls": "2015-16",
    "TSS_2017_FINAL.xlsx":           "2016-17",
    "TSS_2018.xlsx":                 "2017-18",
    "TSS-2019.xlsx":                 "2018-19",
    "TSS-2020.xlsx":                 "2019-20",
    "TSS-2021.xlsx":                 "2020-21",
    "TSS-2022.xlsx":                 "2021-22",
    "TSS-2023.xlsx":                 "2022-23",
    "TSS-2024.xlsx":                 "2023-24",
    "TSS-2025.xlsx":                 "2024-25",
    "TSS-2026.xlsx":                 "2025-26",
}


# ---------------------------------------------------------------------------
# Extended typed columns
# ---------------------------------------------------------------------------
# The full 85-column TSS row is always archived in tss_annual.payload. These
# fields are *promoted* out of the payload into dedicated, queryable columns so
# the AI agent and district profiles can read them without JSON gymnastics.
#
# Each field is resolved by an EXACT (case-insensitive, whitespace-normalized)
# header match against the spreadsheet — robust against the duplicate/truncated
# short headers that a fuzzy substring match would confuse. `kind` selects the
# sanitizer + SQL type:
#   money  -> NUMERIC(12,2), range 0..1,000,000 (premiums/salaries; 0 kept)
#   pct    -> NUMERIC(6,2),  range 0..100 (whole-number percentages)
#   years  -> INTEGER,       range 0..60
#   text   -> TEXT,          trimmed string or NULL
TF = namedtuple("TF", "key column sqltype header kind")

_SQL_BY_KIND = {
    "money": "NUMERIC(12,2)",
    "pct":   "NUMERIC(6,2)",
    "years": "INTEGER",
    "text":  "TEXT",
}

EXTENDED_TSS_FIELDS: list[TF] = [
    # --- Salary schedule structure -----------------------------------------
    TF("salary_program", "salary_program", "TEXT", "Salary Program", "text"),
    TF("ma30_begin", "ma30_begin", "NUMERIC(12,2)",
       "Master's 30-32+ Beginning Salary", "money"),
    TF("ma30_max", "ma30_max", "NUMERIC(12,2)",
       "Master's 30-32+ Maximum Salary", "money"),
    TF("ma30_ytm", "ma30_years_to_max", "INTEGER",
       "Master's 30-32+ Years to Max", "years"),
    TF("hss_ytm", "hss_years_to_max", "INTEGER",
       "Years to Highest Scheduled Salary", "years"),
    TF("education_level_required", "education_level_required", "TEXT",
       "Education Level Required to Reach Highest Salary", "text"),
    TF("masters_10th_year_salary", "masters_10th_year_salary", "NUMERIC(12,2)",
       "MASTERS 10TH YEAR SALARY", "money"),
    # --- Contract / structural flags ---------------------------------------
    TF("severance_pay", "severance_pay", "TEXT", "Severance Pay", "text"),
    TF("early_retirement_program", "early_retirement_program", "TEXT",
       "Early Retirement Program", "text"),
    TF("sick_leave_bank", "sick_leave_bank", "TEXT", "Sick Leave Bank", "text"),
    TF("trs_included_in_salary", "trs_included_in_salary", "TEXT",
       "Teacher Retirement System Monies Included in Salary Schedule", "text"),
    TF("fair_share_provision", "fair_share_provision", "TEXT",
       "DOES CONTRACT CONTAIN FAIR SHARE PROVISION?", "text"),
    TF("longevity_pay_provided", "longevity_pay_provided", "TEXT",
       "LONGEVITY PAY PROVIDED BY CONTRACT?", "text"),
    # --- Benefits matrix: 6 types x {premium, % employer} x {employee, family}
    TF("dental_premium_employee", "dental_premium_employee", "NUMERIC(12,2)",
       "Annual Dental Premium for Employee", "money"),
    TF("dental_pct_employer_employee", "dental_pct_employer_employee", "NUMERIC(6,2)",
       "Percentage of Employee Dental Paid by Employer", "pct"),
    TF("dental_premium_family", "dental_premium_family", "NUMERIC(12,2)",
       "Annual Dental Premium for Family", "money"),
    TF("dental_pct_employer_family", "dental_pct_employer_family", "NUMERIC(6,2)",
       "Percentage of Family Dental Paid by Employer", "pct"),
    TF("disability_premium_employee", "disability_premium_employee", "NUMERIC(12,2)",
       "Annual Disability Insurance Premium for Employee", "money"),
    TF("disability_pct_employer_employee", "disability_pct_employer_employee",
       "NUMERIC(6,2)",
       "Percentage of Employee Disability Insurance Paid by Employer", "pct"),
    TF("disability_premium_family", "disability_premium_family", "NUMERIC(12,2)",
       "Annual Disability Insurance Premium for Family", "money"),
    TF("disability_pct_employer_family", "disability_pct_employer_family",
       "NUMERIC(6,2)",
       "Percentage of Family Disability Insurance Paid by Employer", "pct"),
    TF("health_premium_employee", "health_premium_employee", "NUMERIC(12,2)",
       "Annual Hospitalization Insurance Premium for Employee", "money"),
    TF("health_pct_employer_employee", "health_pct_employer_employee", "NUMERIC(6,2)",
       "Percentage of Employee Hospitalization Insurance Paid by Employer", "pct"),
    TF("health_premium_family", "health_premium_family", "NUMERIC(12,2)",
       "Annual Hospitalization Insurance Premium for Family", "money"),
    TF("health_pct_employer_family", "health_pct_employer_family", "NUMERIC(6,2)",
       "Percentage of Family Hospitalization Insurance Paid by Employer", "pct"),
    TF("life_premium_employee", "life_premium_employee", "NUMERIC(12,2)",
       "Annual Life Insurance Premium for Employee", "money"),
    TF("life_pct_employer_employee", "life_pct_employer_employee", "NUMERIC(6,2)",
       "Percentage of Employee Life Insurance Paid by Employer", "pct"),
    TF("life_premium_family", "life_premium_family", "NUMERIC(12,2)",
       "Annual Life Insurance Premium for Family", "money"),
    TF("life_pct_employer_family", "life_pct_employer_family", "NUMERIC(6,2)",
       "Percentage of Family Life Insurance Paid by Employer", "pct"),
    TF("prescription_premium_employee", "prescription_premium_employee",
       "NUMERIC(12,2)",
       "Annual Prescription Insurance Premium for Employee", "money"),
    TF("prescription_pct_employer_employee", "prescription_pct_employer_employee",
       "NUMERIC(6,2)",
       "Percentage of Employee Prescription Insurance Paid by Employer", "pct"),
    TF("prescription_premium_family", "prescription_premium_family", "NUMERIC(12,2)",
       "Annual Prescription Insurance Premium for Family", "money"),
    TF("prescription_pct_employer_family", "prescription_pct_employer_family",
       "NUMERIC(6,2)",
       "Percentage of Family Prescription Insurance Paid by Employer", "pct"),
    TF("vision_premium_employee", "vision_premium_employee", "NUMERIC(12,2)",
       "Annual Vision Insurance Premium for Employee", "money"),
    TF("vision_pct_employer_employee", "vision_pct_employer_employee", "NUMERIC(6,2)",
       "Percentage of Employee Vision Insurance Paid by Employer", "pct"),
    TF("vision_premium_family", "vision_premium_family", "NUMERIC(12,2)",
       "Annual Vision Insurance Premium for Family", "money"),
    TF("vision_pct_employer_family", "vision_pct_employer_family", "NUMERIC(6,2)",
       "Percentage of Family Vision Insurance Paid by Employer", "pct"),
    # --- Longevity pay: 4 lanes x {max, years-to-max} ----------------------
    TF("longevity_ba_max", "longevity_ba_max", "NUMERIC(12,2)",
       "LONGEVITY BACHELORS MAXIMUM", "money"),
    TF("longevity_ba_ytm", "longevity_ba_years_to_max", "INTEGER",
       "YEARS TO LONGEVITY BACHELORS MAXIMUM", "years"),
    TF("longevity_ma_max", "longevity_ma_max", "NUMERIC(12,2)",
       "LONGEVITY MASTERS MAXIMUM", "money"),
    TF("longevity_ma_ytm", "longevity_ma_years_to_max", "INTEGER",
       "YEARS TO LONGEVITY MASTERS MAXIMUM", "years"),
    TF("longevity_ma30_max", "longevity_ma30_max", "NUMERIC(12,2)",
       "LONGEVITY MASTERS 30-32+ MAXIMUM", "money"),
    TF("longevity_ma30_ytm", "longevity_ma30_years_to_max", "INTEGER",
       "YEARS TO LONGEVITY MASTERS 30-32+ MAXIMUM", "years"),
    TF("longevity_hss_max", "longevity_hss_max", "NUMERIC(12,2)",
       "LONGEVITY HIGHEST SCHEDULED SALARY MAXIMUM", "money"),
    TF("longevity_hss_ytm", "longevity_hss_years_to_max", "INTEGER",
       "LONGEVITY HIGHEST YEAR SCHEDULED SALARY MAXIMUM", "years"),
]

# Base DDL so a fresh database self-creates the table; the ALTER ... ADD COLUMN
# IF NOT EXISTS in ensure_tss_schema() then adds every extended column.
_TSS_BASE_DDL = """
CREATE TABLE IF NOT EXISTS tss_annual (
    id BIGSERIAL PRIMARY KEY,
    state CHAR(2) NOT NULL DEFAULT 'IL',
    state_district_id TEXT NOT NULL,
    school_year VARCHAR(7) NOT NULL,
    district_name TEXT,
    enrollment_range TEXT,
    affiliation TEXT,
    ba_begin NUMERIC(12,2), ba_max NUMERIC(12,2), ba_years_to_max INTEGER,
    ma_begin NUMERIC(12,2), ma_max NUMERIC(12,2), ma_years_to_max INTEGER,
    highest_scheduled_salary NUMERIC(12,2),
    trs_board_paid_pct NUMERIC(6,2),
    contract_expires DATE,
    personal_days NUMERIC(6,1), sick_days NUMERIC(6,1),
    payload JSONB NOT NULL,
    loaded_at TIMESTAMPTZ DEFAULT now(),
    UNIQUE (state, state_district_id, school_year)
);
"""

# Column order for the dynamic INSERT (keys + base typed + extended + payload).
_TSS_BASE_DATA_COLS = [
    "district_name", "enrollment_range", "affiliation",
    "ba_begin", "ba_max", "ba_years_to_max",
    "ma_begin", "ma_max", "ma_years_to_max",
    "highest_scheduled_salary", "trs_board_paid_pct",
    "contract_expires", "personal_days", "sick_days",
]
_TSS_EXT_COLS = [f.column for f in EXTENDED_TSS_FIELDS]
_TSS_INSERT_COLS = (
    ["state", "state_district_id", "school_year"]
    + _TSS_BASE_DATA_COLS + _TSS_EXT_COLS + ["payload"]
)
_TSS_UPDATE_COLS = _TSS_BASE_DATA_COLS + _TSS_EXT_COLS + ["payload"]
_TSS_INSERT_SQL = (
    "INSERT INTO tss_annual (" + ", ".join(_TSS_INSERT_COLS) + ") VALUES ("
    + ", ".join(["%s"] * len(_TSS_INSERT_COLS)) + ") "
    "ON CONFLICT (state, state_district_id, school_year) DO UPDATE SET "
    + ", ".join(f"{c} = EXCLUDED.{c}" for c in _TSS_UPDATE_COLS)
    + ", loaded_at = now()"
)


def _norm_hdr(s: Any) -> str:
    """Normalize a header for exact matching: collapse whitespace, lowercase."""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


# ---------------------------------------------------------------------------
# Slug helper
# ---------------------------------------------------------------------------

def _slug_il(name: str, rcdt: str) -> str:
    """Generate a unique slug for an IL district (prefixed with 'il-').

    Uses the full 11-digit RCDT as a suffix so that two districts with
    identical names in different regions never collide.
    """
    base = "il-" + re.sub(r"[^a-z0-9]+", "-", name.lower().strip())
    digits = re.sub(r"\D", "", rcdt).zfill(11)
    slug = re.sub(r"-+", "-", f"{base}-{digits}").strip("-")
    return slug[:120]  # guard against absurdly long names


# ---------------------------------------------------------------------------
# RCDT normalization
# ---------------------------------------------------------------------------

def _normalize_rcdt(raw: Any) -> Optional[str]:
    """
    Normalize an RCDT value to an 11-digit zero-padded string.
    Returns None for title/footer rows (< 9 or > 11 digits after stripping).
    """
    if raw is None:
        return None
    s = str(raw).strip()
    # Excel often reads numeric codes as floats: "1001001026.0"
    if "." in s:
        s = s.split(".")[0]
    digits = re.sub(r"\D", "", s)
    if not digits or not (9 <= len(digits) <= 11):
        return None
    return digits.zfill(11)


# ---------------------------------------------------------------------------
# Column resolver
# ---------------------------------------------------------------------------

def _col_index(headers: list[str], *candidates: str) -> Optional[int]:
    """
    Return the first column index matching any candidate.
    Short candidates (≤ 5 chars): case-insensitive EXACT match (for short codes
    like BB, BM, LA, Exp, TPBP).
    Longer candidates: case-insensitive SUBSTRING match.
    """
    h_lower = [str(h).lower().strip() for h in headers]
    for cand in candidates:
        cl = cand.lower().strip()
        short = len(cl) <= 5
        for i, h in enumerate(h_lower):
            if short:
                if h == cl:
                    return i
            else:
                if cl in h:
                    return i
    return None


def resolve_columns(headers: list[str]) -> dict[str, Optional[int]]:
    """Map logical field names → column indices (None if absent in this file)."""
    def ci(*cands: str) -> Optional[int]:
        return _col_index(headers, *cands)

    # sick_days: prefer "days sick" prefix over bare "sick" to avoid the yes/no
    # "Sick Leave > 80 Days" column (which does NOT start with "days sick").
    sick = ci("days sick leave", "sl>180#", "sick leave accumulated",
              "sick leave days", "ds")

    # personal_days: PBEL short code or verbose prefix
    personal = ci("days personal", "pbel", "personal, business", "personal days", "dp")
    if personal is None:
        personal = ci("personal")

    col_map = {
        "rcdt":             ci("rcdt code", "rcdt"),
        "name":             ci("district name", "dist name"),
        "enrollment_range": ci("enrollment range", "ere"),
        "affiliation":      ci("local affiliation", "la"),
        "ba_begin":         ci("bachelor's beginning salary", "bachelor's beginning",
                               "bachelor beginning", "bb"),
        "ba_max":           ci("bachelor's maximum salary",   "bachelor's maximum",
                               "bachelor maximum",   "bm"),
        "ba_ytm":           ci("bachelor's years to max", "bytm"),
        "ma_begin":         ci("master's beginning salary", "master's beginning",
                               "master beginning", "mb"),
        "ma_max":           ci("master's maximum salary",   "master's maximum",
                               "master maximum",   "mm"),
        "ma_ytm":           ci("master's years to max", "mytm"),
        "hss":              ci("highest scheduled salary", "hss"),
        "trs_pct":          ci("percentage of board paid trs", "board paid trs",
                               "tpbp", "retirement"),
        "expires":          ci("expiration date of contract", "expiration date", "exp"),
        "personal":         personal,
        "sick":             sick,
    }

    # Extended typed fields: resolve by EXACT normalized header. Absent in older
    # vintages (older files lack these columns) -> None, which is fine; the
    # corresponding typed column is simply left NULL for those rows.
    norm_map = {_norm_hdr(h): i for i, h in enumerate(headers)}
    for f in EXTENDED_TSS_FIELDS:
        col_map[f.key] = norm_map.get(_norm_hdr(f.header))
    return col_map


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------

def _safe_num(val: Any, lo: float, hi: float) -> Optional[float]:
    """Parse val as float; return None if outside [lo, hi] or not parseable."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    try:
        f = float(s.replace(",", ""))
        return f if lo <= f <= hi else None
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any, lo: int, hi: int) -> Optional[int]:
    f = _safe_num(val, lo, hi)
    return int(round(f)) if f is not None else None


def _safe_text(val: Any) -> Optional[str]:
    """Trim to a non-empty string or None (drops blanks / 'nan' / 'none')."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("nan", "none") else None


def _sanitize_ext(kind: str, val: Any) -> Any:
    """Sanitize an extended-field value according to its declared kind."""
    if kind == "money":
        return _safe_num(val, 0, 1_000_000)
    if kind == "pct":
        return _safe_num(val, 0, 100)
    if kind == "years":
        return _safe_int(val, 0, 60)
    if kind == "text":
        return _safe_text(val)
    return None


def _parse_date(val: Any) -> Optional[str]:
    """
    Parse expiration date → 'YYYY-MM-DD' string.
    Handles:
      - Python datetime / date objects (openpyxl returns these for xlsx cells)
      - Excel serial integer (xlrd numeric date cells)
      - Compact MYY / MMYY strings (2015-16 format: "816"=Aug-2016, "1216"=Dec-2016)
      - M/YYYY, M/D/YYYY, YYYY-MM-DD strings
    Returns None for anything outside 2000-2050 or not parseable.
    """
    import datetime as dt
    if val is None:
        return None

    # Python datetime / date objects — openpyxl returns these for date cells
    if isinstance(val, dt.datetime):
        d = val.date()
        return d.isoformat() if 2000 <= d.year <= 2050 else None
    if isinstance(val, dt.date):
        return val.isoformat() if 2000 <= val.year <= 2050 else None

    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None

    # Excel serial number (numeric date from xlrd)
    try:
        serial = float(s)
        if serial != serial:  # NaN guard
            return None
        if 20000 < serial < 60000:
            d = dt.date(1899, 12, 30) + dt.timedelta(days=int(serial))
            return d.isoformat() if 2000 <= d.year <= 2050 else None
    except (ValueError, OverflowError):
        pass

    # M/YYYY
    m = re.match(r"^(\d{1,2})/(\d{4})$", s)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        if 2000 <= year <= 2050 and 1 <= month <= 12:
            return f"{year}-{month:02d}-01"
        return None

    # M/D/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= yr <= 2050 and 1 <= mo <= 12 and 1 <= dy <= 31:
            return f"{yr}-{mo:02d}-{dy:02d}"
        return None

    # YYYY-MM-DD (also catches "YYYY-MM-DD HH:MM:SS" via prefix match)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m and 2000 <= int(m.group(1)) <= 2050:
        yr, mo, dy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= dy <= 31:
            return f"{yr}-{mo:02d}-{dy:02d}"

    # M/YY (two-digit year, slash separator)
    m = re.match(r"^(\d{1,2})/(\d{2})$", s)
    if m:
        mo, yr = int(m.group(1)), 2000 + int(m.group(2))
        if 2000 <= yr <= 2050 and 1 <= mo <= 12:
            return f"{yr}-{mo:02d}-01"

    # Compact MYY or MMYY (no separator — 2015-16 TSS format)
    # e.g. "816" = month 8, year 2016; "1216" = month 12, year 2016
    m = re.match(r"^(\d{1,2})(\d{2})$", s)
    if m:
        mo, yr = int(m.group(1)), 2000 + int(m.group(2))
        if 2000 <= yr <= 2050 and 1 <= mo <= 12:
            return f"{yr}-{mo:02d}-01"

    return None


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def _find_header_row(all_rows: list[list]) -> int:
    """Return the index of the first row containing 'RCDT'."""
    for i, row in enumerate(all_rows):
        if any("RCDT" in str(c).upper() for c in (row or [])
               if c is not None and str(c).strip()):
            return i
    raise ValueError("No row containing 'RCDT' found")


def load_xls(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Load .xls → (headers, data_rows). Finds the correct sheet automatically."""
    import xlrd
    wb = xlrd.open_workbook(str(path))

    target = None
    for idx in range(wb.nsheets):
        sh = wb.sheet_by_index(idx)
        for r in range(min(10, sh.nrows)):
            if any("RCDT" in str(sh.cell_value(r, c)).upper()
                   for c in range(sh.ncols)):
                target = sh
                break
        if target:
            break

    if not target:
        raise ValueError(f"No RCDT sheet found in {path.name}")

    all_rows = []
    for r in range(target.nrows):
        all_rows.append([target.cell_value(r, c) for c in range(target.ncols)])

    hdr_idx = _find_header_row(all_rows)
    headers = [str(v).strip() for v in all_rows[hdr_idx]]
    data_rows = all_rows[hdr_idx + 1:]
    return headers, data_rows


def load_xlsx(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Load .xlsx → (headers, data_rows). Finds the correct sheet automatically."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    target_rows = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[:10]:
            if any("RCDT" in str(c).upper() for c in (row or [])
                   if c is not None and str(c).strip()):
                target_rows = rows
                break
        if target_rows is not None:
            break

    wb.close()

    if target_rows is None:
        raise ValueError(f"No RCDT sheet found in {path.name}")

    hdr_idx = _find_header_row([[c for c in (r or [])] for r in target_rows])
    headers = [str(c).strip() if c is not None else "" for c in target_rows[hdr_idx]]
    data_rows = [list(r or []) for r in target_rows[hdr_idx + 1:]]
    return headers, data_rows


# ---------------------------------------------------------------------------
# Row parser
# ---------------------------------------------------------------------------

def _payload(headers: list[str], row: list[Any]) -> dict:
    """Build a JSON-safe payload dict from the full row."""
    result = {}
    for i, h in enumerate(headers):
        if i >= len(row):
            break
        v = row[i]
        if v is None:
            continue
        # Filter out float NaN
        if isinstance(v, float) and v != v:
            continue
        key = (h or f"col{i}").strip()
        result[key] = v if not isinstance(v, float) else (
            int(v) if v == int(v) and abs(v) < 1e15 else round(v, 6)
        )
    return result


def parse_row(
    row: list[Any],
    col_map: dict[str, Optional[int]],
    headers: list[str],
) -> Optional[dict]:
    """
    Parse one data row into a dict.
    Returns None if the row should be skipped (bad/missing RCDT).
    """
    def get(field: str) -> Any:
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        # Treat float NaN as None
        return None if (isinstance(v, float) and v != v) else v

    rcdt = _normalize_rcdt(get("rcdt"))
    if rcdt is None:
        return None

    name_raw = get("name")
    district_name = str(name_raw).strip() if name_raw is not None else None

    er = get("enrollment_range")
    enrollment_range = str(er).strip() if er is not None and str(er).strip() else None

    aff = get("affiliation")
    affiliation = str(aff).strip() if aff is not None and str(aff).strip() else None

    # Extended typed fields, sanitized per their declared kind.
    ext = {f.key: _sanitize_ext(f.kind, get(f.key)) for f in EXTENDED_TSS_FIELDS}

    return {
        "rcdt":             rcdt,
        "district_name":    district_name,
        "enrollment_range": enrollment_range,
        "affiliation":      affiliation,
        "ba_begin":  _safe_num(get("ba_begin"),  10_000, 500_000),
        "ba_max":    _safe_num(get("ba_max"),     10_000, 500_000),
        "ba_ytm":    _safe_int(get("ba_ytm"),     0, 60),
        "ma_begin":  _safe_num(get("ma_begin"),   10_000, 500_000),
        "ma_max":    _safe_num(get("ma_max"),      10_000, 500_000),
        "ma_ytm":    _safe_int(get("ma_ytm"),      0, 60),
        "hss":       _safe_num(get("hss"),         10_000, 500_000),
        "trs_pct":   _safe_num(get("trs_pct"),     0, 100),
        "expires":   _parse_date(get("expires")),
        "personal":  _safe_num(get("personal"),    0, 100),
        "sick":      _safe_num(get("sick"),         0, 400),
        "ext":       ext,
        "payload":   _payload(headers, row),
    }


# ---------------------------------------------------------------------------
# Database upserts
# ---------------------------------------------------------------------------

def upsert_district(cur, rcdt: str, name: str) -> None:
    slug = _slug_il(name or f"il-district-{rcdt}", rcdt)
    cur.execute(
        """
        INSERT INTO districts (state, state_district_id, name, slug)
        VALUES ('IL', %s, %s, %s)
        ON CONFLICT (state, state_district_id) DO UPDATE SET
            name = EXCLUDED.name,
            slug = EXCLUDED.slug
        """,
        (rcdt, name or f"IL District {rcdt}", slug),
    )


def ensure_tss_schema(conn) -> None:
    """Self-migrate: create tss_annual if missing, then additively add every
    extended typed column. Safe to run repeatedly and in production (no drops,
    no type changes) so an uploaded load applies the schema on its own."""
    cur = conn.cursor()
    cur.execute(_TSS_BASE_DDL)
    for f in EXTENDED_TSS_FIELDS:
        cur.execute(
            f"ALTER TABLE tss_annual ADD COLUMN IF NOT EXISTS {f.column} {f.sqltype}"
        )
    conn.commit()
    cur.close()


def upsert_tss_row(cur, school_year: str, p: dict) -> None:
    """Insert/replace one district-year row, including all extended typed columns.

    The column list and SQL are built once from the declarative spec
    (_TSS_INSERT_SQL); the values tuple below must stay in the same order:
    keys, base data columns, extended columns (spec order), then payload.
    """
    vals: list[Any] = [
        "IL", p["rcdt"], school_year,
        p["district_name"], p["enrollment_range"], p["affiliation"],
        p["ba_begin"], p["ba_max"], p["ba_ytm"],
        p["ma_begin"], p["ma_max"], p["ma_ytm"],
        p["hss"], p["trs_pct"], p["expires"], p["personal"], p["sick"],
    ]
    vals += [p["ext"].get(f.key) for f in EXTENDED_TSS_FIELDS]
    vals.append(json.dumps(p["payload"], default=str))
    cur.execute(_TSS_INSERT_SQL, vals)


# ---------------------------------------------------------------------------
# Per-file loader
# ---------------------------------------------------------------------------

def load_file(conn, path: Path, school_year: str) -> dict:
    """Load one TSS file. Returns stats dict."""
    log.info("%-35s → %s", path.name, school_year)

    # Self-apply the additive schema before loading (idempotent, prod-safe).
    ensure_tss_schema(conn)

    if path.suffix.lower() == ".xls":
        headers, data_rows = load_xls(path)
    else:
        headers, data_rows = load_xlsx(path)

    col_map = resolve_columns(headers)

    # Log which columns were resolved (debug)
    resolved = {k: (headers[v] if v is not None else None) for k, v in col_map.items()}
    log.debug("Column map for %s: %s", path.name, resolved)

    cur = conn.cursor()
    loaded = skipped = 0
    ba_begin_count = expires_count = 0

    for row in data_rows:
        parsed = parse_row(row, col_map, headers)
        if parsed is None:
            skipped += 1
            continue

        rcdt = parsed["rcdt"]
        name = parsed["district_name"] or f"IL District {rcdt}"

        try:
            cur.execute("SAVEPOINT tss_row")
            upsert_district(cur, rcdt, name)
            upsert_tss_row(cur, school_year, parsed)
            cur.execute("RELEASE SAVEPOINT tss_row")
            loaded += 1
            if parsed["ba_begin"] is not None:
                ba_begin_count += 1
            if parsed["expires"] is not None:
                expires_count += 1
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT tss_row")
            log.warning("Row error rcdt=%s year=%s: %s", rcdt, school_year, e)
            skipped += 1

    conn.commit()
    cur.close()
    return {
        "loaded":         loaded,
        "skipped":        skipped,
        "ba_begin_count": ba_begin_count,
        "expires_count":  expires_count,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _load_single(path: Path, school_year: Optional[str]) -> None:
    """Load one uploaded TSS file with an explicit school year (admin upload).

    The school year cannot be auto-detected from TSS data, so it must be given
    on the command line (or fall back to the built-in filename map). Exits
    non-zero on failure so the admin panel surfaces an error.
    """
    if not path.exists():
        log.error("File not found: %s", path)
        sys.exit(1)

    sy = (school_year or "").strip() or FILE_SCHOOL_YEARS.get(path.name)
    if not sy:
        log.error("No school year for %s — pass --school-year YYYY-YY (e.g. 2026-27)",
                  path.name)
        sys.exit(1)
    if not re.match(r"^\d{4}-\d{2}$", sy):
        log.error("Invalid --school-year %r — expected YYYY-YY (e.g. 2026-27)", sy)
        sys.exit(1)

    conn = common.get_db_conn()
    try:
        r = load_file(conn, path, sy)
    except Exception as exc:
        import traceback
        log.error("Fatal error loading %s: %s", path.name, exc)
        traceback.print_exc()
        conn.close()
        sys.exit(1)
    conn.close()

    W = 72
    print()
    print("=" * W)
    print("  Illinois TSS Load Summary (single file)")
    print("=" * W)
    print(f"  File         : {path.name}")
    print(f"  School Year  : {sy}")
    print(f"  Districts    : {r['loaded']:,}")
    print(f"  ba_begin     : {r['ba_begin_count']:,}")
    print(f"  expires      : {r['expires_count']:,}")
    print(f"  Skipped      : {r['skipped']:,}")
    print("=" * W)
    print()

    if r["loaded"] == 0:
        log.error("No rows loaded from %s — the file may not be a TSS export "
                  "(expected an 'RCDT' header row).", path.name)
        sys.exit(1)


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(
        description="Load Illinois Teacher Salary Study (TSS) data into tss_annual",
    )
    parser.add_argument(
        "--file",
        help="Process a single TSS file instead of the built-in file list. "
             "Used by the admin upload.",
    )
    parser.add_argument(
        "--school-year", dest="school_year",
        help="School year in YYYY-YY form (e.g. 2026-27) for the uploaded --file. "
             "Required for files not in the built-in filename map.",
    )
    args = parser.parse_args()

    if args.file:
        _load_single(Path(args.file), args.school_year)
        return

    conn = common.get_db_conn()
    results: list[tuple[str, dict]] = []

    for fname, school_year in FILE_SCHOOL_YEARS.items():
        path = IL_TSS_DIR / fname
        if not path.exists():
            log.warning("File not found, skipping: %s", path)
            results.append((school_year, {"loaded": 0, "skipped": 0,
                                          "ba_begin_count": 0, "expires_count": 0}))
            continue
        try:
            r = load_file(conn, path, school_year)
            results.append((school_year, r))
        except Exception as exc:
            import traceback
            log.error("Fatal error loading %s: %s", fname, exc)
            traceback.print_exc()
            results.append((school_year, {"loaded": 0, "skipped": 0,
                                          "ba_begin_count": 0, "expires_count": 0,
                                          "error": str(exc)}))

    conn.close()

    # Query final totals directly from DB
    conn2 = common.get_db_conn()
    cur = conn2.cursor()
    cur.execute("SELECT COUNT(*) FROM tss_annual WHERE state = 'IL'")
    total_rows: int = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT state_district_id) FROM tss_annual WHERE state = 'IL'")
    distinct_districts: int = cur.fetchone()[0]
    cur.close()
    conn2.close()

    # Summary table
    W = 72
    print()
    print("=" * W)
    print("  Illinois TSS Load Summary")
    print("=" * W)
    print(f"  {'School Year':<12} {'Districts':>10} {'ba_begin':>10} {'expires':>10}  {'Skipped':>8}")
    print(f"  {'-'*12} {'-'*10} {'-'*10} {'-'*10}  {'-'*8}")
    for sy, r in results:
        err = f"  ERROR: {r.get('error','')[:30]}" if r.get("error") else ""
        print(
            f"  {sy:<12} {r['loaded']:>10,} {r['ba_begin_count']:>10,}"
            f" {r['expires_count']:>10,}  {r['skipped']:>8,}{err}"
        )
    print(f"  {'-'*12} {'-'*10} {'-'*10} {'-'*10}  {'-'*8}")
    print()
    print(f"  Total tss_annual rows (IL)   : {total_rows:>10,}")
    print(f"  Distinct IL districts        : {distinct_districts:>10,}")
    print("=" * W)
    print()


if __name__ == "__main__":
    main()
